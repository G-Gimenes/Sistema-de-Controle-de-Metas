from flask import Flask, render_template, request, redirect, session, jsonify
from datetime import datetime
import pandas as pd
import os
import json
from flask_cors import CORS
from threading import Lock
import firebase_admin
from firebase_admin import credentials, auth
from flask import request

app = Flask(__name__)
app.secret_key = 'metasapp'
CORS(app)

# Caminhos
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
EXCEL_FILE = os.path.join(UPLOAD_FOLDER, 'meta_industria.xlsx')
LOG_PATH = os.path.join(BASE_DIR, "logs", "acessos.log")
USERS_PATH = os.path.join(BASE_DIR, "static", "usuarios.json")
os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)

cred = credentials.Certificate(r"Insira seu path/ arquivo_key.json")
firebase_admin.initialize_app(cred)

lock = Lock()

# Funções auxiliares
def carregar_usuarios():
    with open(USERS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def parse_date(val):
    if val is None or str(val).strip() == '':
        return pd.NaT
    val = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return pd.NaT

def verificar_token(token):
    try:
        decoded_token = auth.verify_id_token(token)
        uid = decoded_token['uid']
        return uid
    except Exception as e:
        return None
    
# Rota de login
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["usuario"]
        senha = request.form["senha"]
        try:
            # Login no Firebase
            user = auth.get_user_by_email(email)  # Confere se o usuário existe
            # Para autenticação real (senha), você precisará usar Firebase Client SDK no front
            session["usuario"] = email
            horario = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(LOG_PATH, "a") as f:
                f.write(f"{email} entrou às {horario}\n")
            return redirect("/dashboard")
        except Exception as e:
            return render_template("login_index.html", erro="Usuário ou senha inválidos")

    return render_template("login_index.html")

@app.route("/dashboard")
def dashboard():
    token = request.cookies.get('firebase_token')
    if not token or not verificar_token(token):
        return redirect("/")
    
    try:
        columns = ['Data', 'Unidade', 'Tipo de Meta', 'Meta']
        if not os.path.exists(EXCEL_FILE):
            empty = pd.DataFrame(columns=columns)
            with lock:
                empty.to_excel(EXCEL_FILE, index=False, engine='openpyxl')

        df = pd.read_excel(EXCEL_FILE, dtype=str, engine='openpyxl')
        data = df.fillna('').values.tolist()

        return render_template(
            'index.html',
            data_desossa=data,
            columns_desossa=columns,
            data_abate=[], columns_abate=[]
        )
    except Exception as e:
        return f"<h3>Erro ao carregar arquivo: {e}</h3>"

# Rota para salvar metas
@app.route('/save_metas', methods=['POST'])
def save_metas():
    if "usuario" not in session:
        return jsonify({'error': 'Não autenticado'}), 403

    try:
        payload = request.json
        data = payload.get('data', [])
        columns = ['Data', 'Unidade', 'Tipo de Meta', 'Meta']

        df = pd.DataFrame(data, columns=columns)
        df['Data'] = df['Data'].apply(parse_date)
        df['Meta'] = (
            df['Meta']
            .astype(str)
            .str.replace(',', '.')
            .pipe(pd.to_numeric, errors='coerce')
            .round(2)
        )

        with lock:
            df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
            from openpyxl import load_workbook
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            for idx, col_name in enumerate(columns, start=1):
                if col_name == 'Data':
                    for cell_tuple in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                        for cell in cell_tuple:
                            cell.number_format = "DD/MM/YYYY"
                if col_name == 'Meta':
                    for cell_tuple in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                        for cell in cell_tuple:
                            cell.number_format = "0.00"
            wb.save(EXCEL_FILE)

        return jsonify({'status': 'sucesso'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Logout opcional
@app.route("/logout")
def logout():
    session.pop("usuario", None)
    return redirect("/")

# Inicializa
if __name__ == "__main__":
    app.run(debug=True)